# _*_ coding:utf-8 _*_
# DATE 2021/8/31
import os
import time
import openpyxl
import config
from openpyxl.styles import Font, PatternFill, colors

from common.logger import log_error
from config import dir_config as dir
from config import config
from config.config import casesSheet


class ExcelOperator:

    def __init__(self,excel_path):
        self.filepath=dir.web_testdatas_dir+excel_path
        self.wk = openpyxl.load_workbook(self.filepath)

    def get_sheet_column_value(self,sheetname,row,col):
        """
        获取excel某行某列的具体数据
        :param sheetname:
        :param row:
        :param col:
        :return: 单元格的值
        """
        mysheet=self.wk[sheetname]
        value= mysheet.cell(row,col).value
        return value


    def  get_sheet_col_values(self,col=config.isexecuteCol,sheetname=config.casesSheet):
        """获取工作表某列的值"""
        mysheet = self.wk[sheetname]
        values=[]
        for i in range(1,mysheet.max_row+1):
            value=mysheet.cell(i,col).value
            values.append(str(value))
        return  values

    def get_sheet_startcol_endcol_value(self,sheetname,start=config.keywordCol,end=config.actionCol):
        """
        读取测试用例所在sheet的关键字到操作值区域内所有非空的值，组成一个列表
        :param sheetname:
        :param start:
        :param end:
        :return:
        """
        mysheet = self.wk[sheetname]
        values = []
        for row in range(2, mysheet.max_row + 1):
            stepdata = []
            for col in range(start, end + 1):
                value = mysheet.cell(row, col).value
                # 过滤值为None
                if value != None:
                    stepdata.append(value)
            values.append(stepdata)
        return values


    def get_executecasename(self,sheetname=config.casesSheet):
        """
        统计测试用例汇总中执行的用例
        :param sheetname:
        :return:返回一个列表，列表中是需要执行用例的sheet名称
        """
        sheet=self.wk[sheetname]
        executecasenames=[]
        for row in range(1,sheet.max_row+1):
            if self.get_sheet_column_value(sheetname,row,config.isexecuteCol)=="y":
                casename=self.get_sheet_column_value(sheetname,row,config.casenameCol)
                executecasenames.append(casename)
        return executecasenames


    def  get_executecase_testdata(self):
        """
        对数据驱动的数据进行处理，提取出case的名字和case的内容
        :return:
        """
        casenames = self.get_executecasename()
        # casesheets=self.wk.sheetnames[1:]
        executecasesdata=[]
        for case in casenames:
            # 获取每条用例的操作步骤数据
            casedata={}
            values=self.get_sheet_startcol_endcol_value(case)
            # assert_values=self.get_assert_value(case)
            casedata["casename"]=case
            casedata["casestepdata"]=values
            # casedata["assert"] = assert_values
            executecasesdata.append(casedata)
        return executecasesdata



    def write_test_result(self,sheet_name,row,col=config.stepresultCol,result="FALSE"):
        """
        根据result填写测试结果，FALSE失败是红色，否则就是PASS绿色
        :param sheet_name:
        :param row:
        :param col:
        :param result:
        :return:
        """
        casesheet=self.wk[sheet_name]
        casesheet.cell(row,col).value=result
        #填充颜色
        redfill=PatternFill("solid", fgColor="FF0000")
        greenfill=PatternFill("solid", fgColor="339933")
        if result=="FALSE":
            casesheet.cell(row,col).fill=redfill
            # 写入执行时间
            current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            casesheet.cell(row, col-1).value = current_time
        else:
            casesheet.cell(row, col).fill=greenfill
            current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            casesheet.cell(row, col - 1).value = current_time
        self.wk.save(self.filepath)

    def write_cases_result(self,sheetname=casesSheet,col=config.casesresultCol):
        """
        统计所有测试用例的结果最终填写到测试用例汇总表格中
        :param sheetname:
        :param col:
        :return:
        """
        all_sheet=self.wk[sheetname]
        for row in range(2,all_sheet.max_row+1):
            all_sheet.cell(row,col).value=""
            whitefill = PatternFill("solid", fgColor="00FFFFFF")
            all_sheet.cell(row,col).fill=whitefill
        values=self.get_sheet_startcol_endcol_value(sheetname=sheetname,start=4,end=5)
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        for index,value in enumerate(values):
            if value[1]=="y":
                #获取对应用例的执行结果
                results=self.get_sheet_col_values(config.stepresultCol,value[0])
                if "FALSE" in results:
                    self.write_test_result(sheetname,index+2,col)
                    # 写入执行时间
                    all_sheet.cell(index+2, col-1).value = current_time
                else:
                    self.write_test_result(sheetname,index+2,col,"PASS")
                    all_sheet.cell(index + 2, col - 1).value = current_time
        self.wk.save(self.filepath)


    # def get_sheet_row_values(self,row,sheetname=config.casesSheet):
    #     """获取工作表某行的数据"""
    #     mysheet = self.wk[sheetname]
    #     values=[]
    #     for  col in range(1,mysheet.max_column+1):
    #         value = mysheet.cell(row, col).value
    #         values.append(value)
    #     return values


    # def write_current_time(self, sheet_name, raw_no,col_no=config.runtimeCol):
    #     """用例表格中写入运行时间，当前时间"""
    #     sh = self.wk[sheet_name]
    #     current_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime() )
    #     sh.cell(raw_no, col_no).value = current_time
    #     self.wk.save(self.filepath)
